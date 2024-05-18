from openpyxl import load_workbook

try:
    wb = load_workbook("temperature.xlsx")
    ws = wb.active
    ws["A2"] = 303  # Update cell B2 with the desired value
    wb.save("temperature.xlsx")
    #print("Excel sheet updated successfully.")
except Exception as e:
    print("Error occurred while updating Excel:", e)
