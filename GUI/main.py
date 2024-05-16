import sys
from PyQt5 import QtCore
from PyQt5.QtWidgets import QMainWindow, QApplication, QPushButton, QWidget, QSlider
from PyQt5.QtCore import QTimer
from PyQt5.QtCore import pyqtSlot, QFile, QTextStream
from PyQt5.QtGui import QIcon
from openpyxl import load_workbook
import numpy as np
from motor_ex import motor_program

from sidebar_ui import Ui_MainWindow

class MainWindow(QMainWindow):
    def __init__(self):
        super(MainWindow, self).__init__()

        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)
        
        #Runs the program continuously######################################
        self.timer_5 = QTimer(self)
        self.delay = 500
        self.is_running_5 = False 
        #####################################################################
        
        #To run the python program
        self.ui.start_6.clicked.connect(self.start_motor_program_5)
        ##
        #To stop the program
        self.ui.pushButton_5.clicked.connect(self.stop_motor_program_5)
        ##
        #
        self.ui.load5_slider.sliderPressed.connect(self.stop_continuous_program)
        self.ui.load5_slider.sliderReleased.connect(self.resume_continuous_program)
        self.ui.rpm5_slider.sliderPressed.connect(self.stop_continuous_program)
        self.ui.rpm5_slider.sliderReleased.connect(self.resume_continuous_program)
        self.ui.volt_slider.sliderPressed.connect(self.stop_continuous_program)
        self.ui.volt_slider.sliderReleased.connect(self.resume_continuous_program)
        self.ui.cool_slider.sliderPressed.connect(self.stop_continuous_program)
        self.ui.cool_slider.sliderReleased.connect(self.resume_continuous_program)
        ##
        self.ui.icon_only_widget.hide()
        self.ui.stackedWidget.setCurrentIndex(0)
        #self.ui.vi_btn_2.setChecked(True)

    #Runs the program continuously
    def start_motor_program_5(self):
        if not self.is_running_5:
            self.is_running_5 = True
            self.timer_5.timeout.connect(self.ui.run_motor_program)
            speed = self.ui.get_speed_value()
            load = self.ui.get_load_value()
            voltage = self.ui.get_voltage_value()
            coolant_flow = self.ui.get_coolant_flow_value()
            motor_result = motor_program(speed, load, voltage, coolant_flow)
            self.ui.load5_slider.setEnabled(True)
            self.ui.rpm5_slider.setEnabled(True)
            self.ui.volt_slider.setEnabled(True)
            self.ui.cool_slider.setEnabled(True)
            self.ui.start_6.setEnabled(False)
            self.ui.motor_select_6.setEnabled(False)
            self.timer_5.start(self.delay)
    ###
    #Stops the continuously running program
    def stop_motor_program_5(self):
        if self.is_running_5:
            self.is_running_5 = False
            self.timer_5.stop()
            self.ui.load5_slider.setEnabled(False)
            self.ui.rpm5_slider.setEnabled(False)
            self.ui.volt_slider.setEnabled(False)
            self.ui.cool_slider.setEnabled(False)
            self.ui.start_6.setEnabled(True)
            self.ui.motor_select_6.setEnabled(True)
            wb = load_workbook("temperature.xlsx")
            ws = wb.active
            ws["A2"] = 303  
            wb.save("temperature.xlsx")
            # initialization
            motor_current_A = np.array([])
            motor_voltage_V = np.array([])
            motor_input_power_W =np.array([])
            motor_eff_percentage = np.array([])
            motor_output_power_W = np.array([])
            motor_power_loss_W = np.array([])
            motor_torque_out_Nm = np.array([])
            motor_regen_power_W = np.array([])
            motor_surface_temperature_K = np.array([])
            motor_surface_change_in_temp_K = np.array([])
            motor_coolant_flow_mps = np.array([])
    ##########################################################
    # Function to stop the program when slider is pressed
    def stop_continuous_program(self):
        if self.is_running_5:
            self.is_running_5 = False
            self.timer_5.stop()

    # Function to resume the program when slider is released
    def resume_continuous_program(self):
        if not self.is_running_5:
            self.is_running_5 = True
            self.timer_5.start(self.delay)

if __name__ == "__main__":
    app = QApplication(sys.argv)

    with open("style.qss", "r") as style_file:
        style_str = style_file.read()
    app.setStyleSheet(style_str)

    window = MainWindow()
    window.show()

    sys.exit(app.exec())